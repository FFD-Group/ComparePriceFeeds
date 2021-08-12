from feed import Feed
from datetime import date
from openpyxl import Workbook

class FeedReport:

    def __init__(self, filename=None) -> None:
        self.wb = Workbook()
        today = date.today().strftime("%d-%m-%y")
        self.report_sheet_name = "Report-" + today
        self.filename = filename if filename else self.report_sheet_name + ".xlsx"
        self.wb.save(self.filename)
        self.wb["Sheet"].title = self.report_sheet_name

    def write_newly_oos(self, items: list, add_due_dates: bool=False):
        oos_sh = self.wb.create_sheet("Newly OOS")
        oos_sh.append(['sku', 'is_in_stock'])
        if add_due_dates:
            dd_sh = self.wb.create_sheet("Due Dates")
            dd_sh.append(['sku', 'due date'])
        for product in items:
            if add_due_dates:
                oos_sh.append([product[0], 0])
                dd_sh.append(product)
            else:
                oos_sh.append([product, 0])

        report_sh = self.wb[self.report_sheet_name]
        report_sh.append(["OOS Items:", len(items)])
        self.wb.save(self.filename)

    def write_back_in(self, items: list):
        bis_sh = self.wb.create_sheet("Back in")
        bis_sh.append(['sku', 'is_in_stock'])
        for product in items:
            if len(product) < 2:
                bis_sh.append([product, 1])
            else:
                bis_sh.append([product[0], 1])
        
        report_sh = self.wb[self.report_sheet_name]
        report_sh.append(["Back in stock:", len(items)])
        self.wb.save(self.filename)

    def write_prices(self, items: list, sheetname: str, f: Feed):
        if sheetname not in self.wb.sheetnames:
            sh = self.wb.create_sheet(sheetname)
        else:
            sh = self.wb[sheetname]
        sh.append(["sku", "price", "special_price", "cost"])
        for sku, diff in items:
            price = f.get_product_data(sku, "RRP", "ProductCode")
            cost = f.get_product_data(sku, "WebOnlyPrice", "ProductCode")
            special_price = self.calculate_markup(cost)
            sh.append([sku, price, special_price, cost])
        self.wb.save(self.filename)

    def calculate_markup(self, cost) -> float:
        if cost <= 9.99:
            #30%
            return cost * 1.30
        elif cost <= 399.99:
            #15%
            return cost * 1.15
        else:
            #10%
            return cost * 1.10

    def write_report_sheet(self, rows, sheetname):
        if sheetname not in self.wb.sheetnames:
            report_sh = self.wb.create_sheet(sheetname)
        else:
            report_sh = self.wb[sheetname]
        for row in rows:
            report_sh.append(row)
        self.wb.move_sheet(sheetname, -(len(self.wb.sheetnames)-1))
        self.wb.save(self.filename)

    def write_dropped(self, items: list):
        dl_sh = self.wb.create_sheet("Dropped")
        dl_sh.append(["sku", "discontinued", "is_in_stock"])
        for product in items:
            dl_sh.append([product, 1, 0])

        report_sh = self.wb[self.report_sheet_name]
        report_sh.append(["Dropped Items:", len(items)])
        self.wb.save(self.filename)


    